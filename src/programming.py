import numpy as np
from openpyxl import load_workbook
import os

app_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
workbook_path = app_dir + '/data/skew.xlsx'
workbook = load_workbook(filename=workbook_path)


def prepare_raw_data(structure, load_limit, include_displacement_limit=False):

    # possible minmax are:
    # minimization: min, maximization: max

    minmax = "max"

    # analysis_type = "static"
    phi = structure.phi
    p0 = structure.p0
    pv = structure.pv
    yield_points_pieces = structure.yield_points_pieces

    phi_pv_phi = phi.T * pv * phi
    phi_p0 = phi.T * p0

    extra_numbers_num = 2 if include_displacement_limit else 1
    total_yield_pieces_num = phi.shape[1]
    variables_num = extra_numbers_num + total_yield_pieces_num

    empty_a = np.zeros((variables_num, variables_num))
    raw_a = np.matrix(empty_a)

    # np.savetxt("phiPvPhi.csv", phi_pv_phi, delimiter=",")
    # np.savetxt("phiP0.csv", phi_p0, delimiter=",")

    raw_a[0:total_yield_pieces_num, 0:total_yield_pieces_num] = phi_pv_phi[0:total_yield_pieces_num, 0:total_yield_pieces_num]
    raw_a[0:total_yield_pieces_num, total_yield_pieces_num] = phi_p0[0:total_yield_pieces_num, 0]
    raw_a[total_yield_pieces_num, total_yield_pieces_num] = 1.0
    b = np.ones((variables_num))

    # if analysis_type == "dynamic":
    #     pass
    #     b[0:-extra_numbers_num] = b[0:-extra_numbers_num] - np.dot(phi_pv_phi, xn_previous[0:-extra_numbers_num])
    # elif analysis_type == "static":
    #     b[0:-extra_numbers_num] = b[0:-extra_numbers_num]

    b[-extra_numbers_num] = load_limit

    # possible inequality_condition are:
    # lt: Less Than or Equal  gt: Larger Than or Eqaul  eq: Equal
    inequality_condition = np.full((variables_num), "lt")

    c = np.zeros(2 * variables_num)
    c[0:total_yield_pieces_num] = 1.0
    mp_data = {
        "variables_num": variables_num,
        "extra_numbers_num": extra_numbers_num,
        "raw_a": raw_a,
        "b": b,
        "c": c,
        "minmax": minmax,
        "inequality_condition": inequality_condition,
        "yield_points_pieces": yield_points_pieces
    }
    return mp_data


def _zero_out_small_values(matrix, floor=1e-8):
    # TODO: why generate small numbers. Do not consider this function.
    # where values are low
    low_values_flags = abs(matrix) < floor
    # all low values set to 0
    matrix[low_values_flags] = 0
    return matrix


def _find_pivot_element(basic_variables, tableau):
    variables_num = basic_variables.shape[0]
    for i in range(variables_num):
        if tableau[i, variables_num] > 0.0:
            min_ba = tableau[i, variables_num + 3] / tableau[i, variables_num]
    for i in range(variables_num):
        if tableau[i, variables_num] > 0.0:
            ba = tableau[i, variables_num + 3] / tableau[i, variables_num]
            if ba <= min_ba:
                min_ba = ba
                will_out_row_num = i
                will_out = basic_variables[will_out_row_num]
    return will_out_row_num, will_out


def _pivot_operation_on_pivot_element(will_out_row_num, will_in, basic_variables, tableau):
    variables_num = basic_variables.shape[0]
    basic_variables[will_out_row_num] = will_in
    pivot_element = tableau[will_out_row_num, variables_num]
    all_columns_num = variables_num + 4
    all_rows_num = variables_num + 1

    # Pivot Operation on pivot row
    for i in range(all_columns_num):
        tableau[will_out_row_num, i] = tableau[will_out_row_num, i] / pivot_element

    tableau = _zero_out_small_values(tableau)

    # Pivot Operation on other rows
    for i in range(all_rows_num):
        if i != will_out_row_num:
            pivot_columnt_i_member = tableau[i, variables_num]
            for j in range(all_columns_num):
                tableau[i, j] = tableau[i, j] - pivot_columnt_i_member * tableau[will_out_row_num, j]

    tableau = _zero_out_small_values(tableau)
    return basic_variables, tableau


def complementarity_programming(mp_data):
    # CA==1: hameye C haye manfi daraye a namosbat hastand
    # CA==2: hameye C haye manfi daraye a namosbat nistand
    # nC: No. of Constraint Equations
    # nV: No. of Variables
    # minmax: A parameter which declares that optimization is minimization of maximization
    # 1: Minimization   2: Maximization
    # equ: is a parameter to show the constraint is equality or unequality
    # 1: Less Than or Equal  2: Larger Than or Eqaul  3: Equal
    # A: Matrix of Coefficient of Variables
    # b: Vector of Constants
    # C: Vector of Cost Factors
    # f: Value of optimization
    # dn: Parameter to check that d is nonnegative or not. 1: Negative, 2: Nonnegative
    # ba: bi/ais
    # minba: minimume of bi/ais
    # bV: Basic Variable no.
    # !jj: A Parameter that in each iteration if xy==0 its value will increase by 1. At the first jj=0.
    # !xy: A parameter to distinguish wheather if x is in basis, y=xs or not? xy==0: x in basis and y=xs,
    # vice versa. xy==1: x in basis and y/=Xs, vice versa.
    variables_num = mp_data["variables_num"]
    raw_a = mp_data["raw_a"]
    b = mp_data["b"]
    c = mp_data["c"]
    minmax = mp_data["minmax"]
    inequality_condition = mp_data["inequality_condition"]
    extra_numbers_num = mp_data["extra_numbers_num"]

    cbar = np.zeros((2 * variables_num))

    # ------------------------------------Construction of Aj
    # Apply Inequality Conditions
    for i in range(variables_num):
        if inequality_condition[i] == "gt":
            b[i] = -b[i]
            raw_a[i, 0:variables_num] = -raw_a[i, 0:variables_num]

    # -------------Creating Minimization Objective Function
    if minmax == "max":  # Maximization
        c[0:variables_num] = -c[0:variables_num]
    # ----------------------------Creation of Aj Matrix
    # the first [1:nC,1:nV] arrays of Aj are arrays of A, and the others except the main diagonal arrays that are (1), are zero.
    a = np.zeros((variables_num, 2 * variables_num))
    a[0:variables_num, 0:variables_num] = raw_a[0:variables_num, 0:variables_num]
    j = variables_num

    # Assigning diagonal arrays of y variables.
    for i in range(variables_num):
        a[i, j] = 1.0
        j += 1

    # excel prints
    for i in range(a.shape[0]):
        for j in range(a.shape[1]):
            workbook['table'].cell(row=i + 2, column=j + 2).value = a[i, j]

    for i in range(b.shape[0]):
        workbook['table'].cell(row=i + 2, column=a.shape[1] + 2).value = b[i]
    workbook.save(filename=workbook_path)

    for i in range(c.shape[0]):
        workbook['table'].cell(row=a.shape[0] + 2, column=i + 2).value = c[i]
    workbook.save(filename=workbook_path)

    # In dynamic analysis we may have negative b, so to achieve canonical form we must use 2phase mathematical programming.
    # So that we must check b values and if we had negative ones, we must use 2phase programming.
    # So at the first we check if we have negative b or not:
    # sorted_b = np.sort(b)
    # sorted_b_indices = np.argsort(b)

    # set the initial basic variables
    basic_variables = np.zeros(variables_num)
    for i in range(variables_num):
        basic_variables[i] = variables_num + i

    # if np.any(b < 0):
    negative_constraints_num = 0
    negative_constraints = []
    for i in range(variables_num):
        if b[i] < 0:
            b[i] = -b[i]
            a[i, 0:variables_num] = -a[i, 0:variables_num]
            basic_variables[i] = i + 2 * variables_num
            negative_constraints_num += 1
            negative_constraints.append(i)

    a2phase = np.zeros((variables_num, 2 * variables_num + negative_constraints_num))
    a2phase[0:variables_num, 0:2 * variables_num] = a[0:variables_num, 0:2 * variables_num]
    j = 0
    # Assigning diagonal arrays of z variables.
    for i in range(len(negative_constraints)):
        a2phase[negative_constraints[i], 2 * variables_num + j] = 1
        j += 1
    # np.savetxt("Aj2Phase.csv", aj2phase, delimiter=",")

    # # Computation of vector d
    # d = np.zeros((2 * variables_num + negative_constraints_num))
    # d[0:2 * variables_num] = -np.sum(a[0:variables_num, 0:2 * variables_num], axis=0)

    tableau = np.zeros((variables_num + 2, variables_num + 4))

    j = 0
    # Assigning diagonal arrays of y variables.
    for i in range(variables_num):
        tableau[i, j] = 1.0
        j += 1

    tableau[0:variables_num, variables_num + 3] = b
    tableau[variables_num, variables_num + 1] = 1.0
    tableau[variables_num + 1, variables_num + 2] = 1.0
    tableau[variables_num + 1, variables_num + 3] = -np.sum(b)
    tableau = _zero_out_small_values(tableau)

    # dbar = np.zeros(2 * variables_num + negative_constraints_num)
    # dbar = np.dot(tableau[variables_num + 1, 0:variables_num], a2phase[0:variables_num, 0:2 * variables_num + negative_constraints_num]) + d
    # dbar = _zero_out_small_values(dbar)

    # At the first the entering element is Loading multiplier (Lambda)
    will_in = variables_num - 1

    # Calculation of Abars
    tableau[0:variables_num, variables_num] = a[0:variables_num, will_in]

    # Finding the exiting variable (r)
    will_out_row_num, will_out = _find_pivot_element(basic_variables, tableau)

    # Pivot Operation
    basic_variables, tableau = _pivot_operation_on_pivot_element(will_out_row_num, will_in, basic_variables, tableau)
    cbar[0:2 * variables_num] = np.dot(tableau[variables_num, 0:variables_num], a[0:variables_num, 0:2 * variables_num]) + c[0:2 * variables_num]
    cbar = _zero_out_small_values(cbar)
    will_in = int(will_out - variables_num)
    index_ca = np.argsort(cbar)
    atemp = np.dot(tableau[0:variables_num, 0:variables_num], a[0:variables_num, index_ca[will_in]])
    atemp = _zero_out_small_values(atemp)

    #  -----------------------------Checking if all Abars are non-positive or not
    for i, ai in enumerate(atemp):
        if ai > 0 and basic_variables[i] != variables_num - 1 and basic_variables[i] != 2 * variables_num - 1:
            unbounded = False
            break
        elif ai <= 0 and basic_variables[i] != variables_num - 1 and basic_variables[i] != 2 * variables_num - 1:
            unbounded = True
    # while unbounded is True:
    #     if atemp[i] > 0 and bV[i] != nC-1 and bV[i] != nC+nV-1:
    #         unbounded = False
    #     elif atemp[i] <= 0 and bV[i] != nC-1 and bV[i] != nC+nV-1:
    #         unbounded = True
    #     i += 1
    while will_out != 2 * variables_num - 1 and unbounded is False:
        # Calculation of Abars
        # tableau[0:nC, nC] = Aj[0:nC, s]
        tableau[0:variables_num, variables_num] = np.dot(tableau[0:variables_num, 0:variables_num], a[0:variables_num, will_in])
        tableau[variables_num, variables_num] = cbar[will_in]
        tableau = _zero_out_small_values(tableau)
        will_out_row_num, will_out = _find_pivot_element(basic_variables, tableau)
        # Pivot Operation
        basic_variables, tableau = _pivot_operation_on_pivot_element(will_out_row_num, will_in, basic_variables, tableau)
        cbar[0:2 * variables_num] = c[0:2 * variables_num] + np.dot(tableau[variables_num, 0:variables_num], a[0:variables_num, 0:2 * variables_num])
        cbar = _zero_out_small_values(cbar)
        will_in = int(will_out - variables_num)
        index_ca = np.argsort(cbar)
        atemp = np.dot(tableau[0:variables_num, 0:variables_num], a[0:variables_num, index_ca[will_in]])
        atemp = _zero_out_small_values(atemp)
        # unbounded = True
        # i = 0
        #  -----------------------------Checking if all Abars are non-positive or not
        for i, ai in enumerate(atemp):
            if ai > 0 and basic_variables[i] != variables_num - 1 and basic_variables[i] != 2 * variables_num - 1:
                unbounded = False
                break
            elif ai <= 0 and basic_variables[i] != variables_num - 1 and basic_variables[i] != 2 * variables_num - 1:
                unbounded = True
        # while unbounded is True:
        #     print(i)
        #     print(atemp[i])
        #     if atemp[i] > 0 and bV[i] != nC-1 and bV[i] != nC+nV-1:
        #         unbounded = False
        #     elif atemp[i] <= 0 and bV[i] != nC-1 and bV[i] != nC+nV-1:
        #         unbounded = True
        #     i += 1
    empty_xn = np.zeros((variables_num, 1))
    xn = np.matrix(empty_xn)
    for i in range(variables_num):
        if int(basic_variables[i]) < variables_num:
            xn[int(basic_variables[i]), 0] = tableau[i, variables_num + 3]
    plastic_multipliers = xn[0:-extra_numbers_num, 0]
    return plastic_multipliers


def get_will_in(fpm):
    will_in = fpm
    return will_in


def calculate_abar(full_a_matrix, will_in, b_matrix_inv):
    a = full_a_matrix[:, will_in]
    abar = np.dot(b_matrix_inv, a)
    return abar


def calculate_bbar(b, b_matrix_inv):
    bbar = np.dot(b_matrix_inv, b)
    return bbar


def get_will_out(abar, bbar, basic_variables):
    # TODO: we check b/a to be positive, correct way is to check a to be positive
    # b is not always positive
    # TODO: exclude load variable
    # TODO: do not divide zero values
    # ba = np.round(ba, 5)
    ba = bbar / abar
    minba = min(ba[ba > 0])
    will_out_row_num = np.where(ba == minba)[0][0]
    will_out = basic_variables[will_out_row_num]
    return will_out, will_out_row_num


def get_initial_basic_variables(variables_num):
    basic_variables = np.zeros(variables_num, dtype=int)
    for i in range(variables_num):
        basic_variables[i] = variables_num + i
    return basic_variables


def get_full_a_matrix(variables_num, a_matrix):
    full_a_matrix = np.zeros((variables_num, 2 * variables_num))
    full_a_matrix[0:variables_num, 0:variables_num] = a_matrix[0:variables_num, 0:variables_num]
    j = variables_num

    # Assigning diagonal arrays of y variables.
    for i in range(variables_num):
        full_a_matrix[i, j] = 1.0
        j += 1

    return full_a_matrix


def get_fpm(will_out, variables_num):
    # TODO: check wether is possible to a x be will_out or not
    fpm = will_out - variables_num
    return int(fpm)


def update_b_matrix_inverse(b_matrix_inv, abar, will_out_row_num, variables_num):
    e = np.eye(variables_num)
    eta = np.zeros(variables_num)
    will_out_item = abar[will_out_row_num]

    for i, item in enumerate(abar):
        if i == will_out_row_num:
            eta[i] = 1 / will_out_item
        else:
            eta[i] = -item / will_out_item
    e[:, will_out_row_num] = eta
    updated_b_matrix_inv = np.dot(e, b_matrix_inv)
    return updated_b_matrix_inv


def get_min_cost_variable_num(entering_candidates):
    candidates_costs = [candidate["variable_cost"] for candidate in entering_candidates]
    min_cost_candidate_num = min(range(len(candidates_costs)), key=candidates_costs.__getitem__)
    min_cost_variable_num = entering_candidates[min_cost_candidate_num]["variable_num"]
    return min_cost_variable_num


def is_variable_plastic_multiplier(variable_num, variables_num, extra_numbers_num):
    return False if variable_num >= variables_num - extra_numbers_num else True


def is_candidate_fpm(min_cost_variable_num, variables_num, extra_numbers_num):
    # fpm: free plastic multiplier
    if is_variable_plastic_multiplier(min_cost_variable_num, variables_num, extra_numbers_num) or min_cost_variable_num == variables_num - extra_numbers_num:
        return True
    else:
        return False


def is_will_out_opm(will_out, variables_num, extra_numbers_num):
    # opm: obstacle plastic multiplier
    return is_variable_plastic_multiplier(will_out, variables_num, extra_numbers_num)


def get_yield_point_num_from_piece(piece, yield_points_pieces):
    for yield_point_num, yield_point_pieces in enumerate(yield_points_pieces):
        if piece in yield_point_pieces:
            return yield_point_num


def get_active_yield_points(basic_variables, yield_points_pieces):
    active_yield_points = []
    for variable in basic_variables:
        active_yield_point = get_yield_point_num_from_piece(variable, yield_points_pieces)
        if active_yield_point is not None:
            active_yield_points.append(active_yield_point)
    return active_yield_points


def is_fpm_for_an_active_yield_point(fpm, active_yield_points):
    return True if fpm in active_yield_points else False


def update_basic_variables(basic_variables, will_out_row_num, will_in):
    basic_variables[will_out_row_num] = will_in
    return basic_variables


def reset(basic_variables, b_history, b):
    # INCOMPLETE:
    for basic_variable in basic_variables:
        b_history += b
        b = 0
    return basic_variables, b


def update_entering_candidates(
        entering_candidates,
        old_fpm,
        will_out_row_num,
        cbar,
        variables_num,
        extra_numbers_num,
        basic_variables):

    for candidate in entering_candidates:
        if candidate["variable_num"] == old_fpm:
            entering_candidates.remove(candidate)
            break

    # TODO: check when test opm
    if basic_variables[will_out_row_num] >= variables_num:
        new_fpm = basic_variables[will_out_row_num] - variables_num
    else:
        new_fpm = old_fpm

    new_fpm_candidate = {
        "variable_num": new_fpm,
        "variable_cost": cbar[new_fpm],
    }

    if old_fpm != variables_num - extra_numbers_num:
        old_fpm_slack_num = variables_num + old_fpm
        old_fpm_slack_candidate = {
            "variable_num": old_fpm_slack_num,
            "variable_cost": cbar[old_fpm_slack_num],
        }
        entering_candidates.append(old_fpm_slack_candidate)

    entering_candidates.append(new_fpm_candidate)
    return new_fpm, entering_candidates


def solve_by_mahini_approach(mp_data):

    variables_num = mp_data["variables_num"]
    a_matrix = np.array(mp_data["raw_a"])
    b = mp_data["b"]
    c = -1 * mp_data["c"]
    extra_numbers_num = mp_data["extra_numbers_num"]
    yield_points_pieces = mp_data["yield_points_pieces"]

    full_a_matrix = get_full_a_matrix(variables_num, a_matrix)
    basic_variables = get_initial_basic_variables(variables_num)
    active_yield_points = []
    b_matrix_inv = np.eye(variables_num)
    cb = np.zeros(variables_num)
    fpm = variables_num - extra_numbers_num

    entering_candidates = [
        {
            "variable_num": fpm,
            "variable_cost": 0,
        }
    ]

    landa_bar_var_num = 2 * variables_num - extra_numbers_num
    will_out = 0

    while will_out != landa_bar_var_num:
        min_cost_variable_num = get_min_cost_variable_num(entering_candidates)

        if is_candidate_fpm(min_cost_variable_num, variables_num, extra_numbers_num):
            will_in = get_will_in(fpm)
        else:
            pass

        abar = calculate_abar(full_a_matrix, will_in, b_matrix_inv)
        bbar = calculate_bbar(b, b_matrix_inv)
        will_out, will_out_row_num = get_will_out(abar, bbar, basic_variables)

        if is_will_out_opm(will_out, variables_num, extra_numbers_num):
            pass
        else:
            if is_fpm_for_an_active_yield_point(fpm, active_yield_points):
                pass

        pi_transpose = np.dot(cb, b_matrix_inv)
        cbar = np.zeros(2 * variables_num)

        for i in range(2 * variables_num):
            cbar[i] = c[i] - np.dot(pi_transpose, full_a_matrix[:, i])

        cb[will_out_row_num] = c[will_in]

        fpm, entering_candidates = update_entering_candidates(
            entering_candidates,
            fpm,
            will_out_row_num,
            cbar,
            variables_num,
            extra_numbers_num,
            basic_variables,
        )

        basic_variables = update_basic_variables(basic_variables, will_out_row_num, will_in)
        active_yield_points = get_active_yield_points(basic_variables, yield_points_pieces)
        b_matrix_inv = update_b_matrix_inverse(b_matrix_inv, abar, will_out_row_num, variables_num)

    bbar = np.dot(b_matrix_inv, b)

    empty_xn = np.zeros((variables_num, 1))
    xn = np.matrix(empty_xn)
    for i in range(variables_num):
        if int(basic_variables[i]) < variables_num:
            xn[int(basic_variables[i]), 0] = bbar[i]
    plastic_multipliers = xn[0:-extra_numbers_num, 0]

    return plastic_multipliers
